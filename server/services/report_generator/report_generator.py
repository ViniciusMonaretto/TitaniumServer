import tempfile

from openpyxl.cell import WriteOnlyCell
from services.sensor_data_storage.sensor_data_storage import SensorDataStorage
from services.report_generator.report_generator_commands import ReportGeneratorCommands
from services.config_handler.config_handler import ConfigHandler
from support.logger import Logger
from middleware.client_middleware import ClientMiddleware
from ..service_interface import ServiceInterface
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from typing import Dict, Tuple


class ReportGenerator(ServiceInterface):
    _sensor_data_storage: SensorDataStorage
    _config_handler: ConfigHandler

    def __init__(self, middleware: ClientMiddleware, sensor_data_storage: SensorDataStorage, config_handler: ConfigHandler):
        self._logger = Logger()
        self._middleware = middleware
        self._sensor_data_storage = sensor_data_storage
        self._config_handler = config_handler
        self.initialize_commands()

        self._logger .info("Report Generator initialized")

    def _get_topic_mappings(self) -> Tuple[Dict[str, str], Dict[str, str]]:
        """Get topic to name and type mappings from ConfigHandler."""
        return self._config_handler.get_topic_mappings()

    def initialize_commands(self):
        commands = {
            ReportGeneratorCommands.GENERATE_REPORT: self.generate_report_command,
        }
        self._middleware.add_commands(commands)

    def generate_report_command(self, command):
        """
        Command to generate a report based on the provided data.
        """
        self._logger.info(
            f"Generating report with command: {command}")

        # Call read_sensor_info with a callback to handle the response
        self._sensor_data_storage.read_sensor_info(
            command, self._handle_sensor_data_for_excel)

    def _handle_sensor_data_for_excel(self, result, data_out):
        """
        Callback to handle sensor data and create Excel report.
        """
        if result and data_out and 'info' in data_out:
            try:
                excel_file_path = self.create_excel_report_fast(
                    data_out['info'])
                self._logger.info(
                    f"Excel report created successfully: {excel_file_path}")
                self._middleware.send_command_answear(
                    True,
                    {"status": "success", "file_path": excel_file_path,
                        "report": data_out['info']},
                    data_out["commandId"]
                )
            except Exception as e:
                self._logger.error(f"Failed to create Excel report: {e}")
                self._middleware.send_command_answear(
                    False,
                    {"status": "error", "message": f"Erro ao criar Relatório excel: {e}"},
                    data_out["commandId"]
                )
        else:
            self._logger.error("Failed to generate report - no data received")
            self._middleware.send_command_answear(
                False,
                {"status": "error",
                    "message": "Failed to generate report - no data received"},
                data_out["commandId"]
            )

    def create_excel_report_fast(self, sensor_data, output_dir: str = None):
        """
        Generate a full Excel report with:
        - Main sensor data sheet
        - Power summary sheet
        """

        topic_to_name, topic_to_type = self._get_topic_mappings()

        if not sensor_data:
            raise Exception("Não há dados para criar Relatório excel")

        # Use temporary directory if none provided
        if output_dir is None:
            output_dir = tempfile.mkdtemp()

        # -------------------------------
        # 1. Collect unique timestamps as strings and sort
        # -------------------------------
        timestamps = sorted({
            entry["timestamp"]
            for entries in sensor_data.values()
            for entry in entries
        })

        # -------------------------------
        # 2. Determine topics in order
        # -------------------------------
        topics_in_order = [t for t in sensor_data.keys() if t in topic_to_name]

        # -------------------------------
        # 3. Pre-format timestamps (ISO 8601 -> readable)
        # -------------------------------
        formatted_ts_map = {}
        for ts in timestamps:
            try:
                ts_str = ts.replace('Z', '+00:00')
                dt = datetime.fromisoformat(ts_str)
                formatted_ts_map[ts] = dt.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                formatted_ts_map[ts] = ts

        # -------------------------------
        # 4. Build lookup dict for fast access
        # -------------------------------
        sensor_lookup = {
            topic: {entry["timestamp"]: entry["value"] for entry in entries}
            for topic, entries in sensor_data.items()
        }

        # --- Style settings ---
        bold = Font(bold=True)
        fill_header = PatternFill(
            start_color="b7e1cd", end_color="b7e1cd", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        align_left = Alignment(horizontal="left")
        align_center = Alignment(horizontal="center")

        def styled_row(ws, values, center=False):
            """Return a list of styled WriteOnlyCell for write_only worksheet"""
            row_cells = []
            for v in values:
                cell = WriteOnlyCell(ws, value=v)
                cell.font = bold
                cell.fill = fill_header
                cell.border = border
                cell.alignment = align_center if center else align_left
                row_cells.append(cell)
            return row_cells

        # -------------------------------
        # 5. Create workbook (write_only)
        # -------------------------------
        wb = Workbook(write_only=True)

        # --- Sensor Data Sheet ---
        ws_main = wb.create_sheet("Dados dos Sensores")

        # Append styled header row
        ws_main.append(styled_row(
            ws_main, ["Grupo:", "", "", "", ""]))
        ws_main.append(styled_row(ws_main, [
                       "Timestamp"] + [topic_to_name.get(t, t) for t in topics_in_order], center=True))

        # Append data rows (no style for performance)
        for ts in timestamps:
            row = [formatted_ts_map[ts]]
            for topic in topics_in_order:
                row.append(sensor_lookup[topic].get(ts))
            ws_main.append(row)

        # --- Power Summary Sheet ---
        ws_power = wb.create_sheet("Resumo de Energia")

        # Identify electrical sensors
        electrical_types = ["Power", "Current", "Tension", "PowerFactor"]
        electrical_topics = [
            t for t in topics_in_order if topic_to_type.get(t) in electrical_types]
        electrical_names = [topic_to_name.get(t, t) for t in electrical_topics]

        # Only create Power Summary sheet if there is at least one electrical sensor
        if electrical_topics:
            electrical_names = [topic_to_name.get(
                t, t) for t in electrical_topics]

            # Compute statistics
            total_power = 0.0
            pf_values, current_values = [], []

            for topic in electrical_topics:
                ttype = topic_to_type.get(topic)
                for entry in sensor_data[topic]:
                    value = float(entry["value"])
                    if ttype == "Power":
                        total_power += value
                    elif ttype == "PowerFactor":
                        pf_values.append(value)
                    elif ttype == "Current":
                        current_values.append(value)

            avg_pf = sum(pf_values) / len(pf_values) if pf_values else 0
            max_current = max(current_values) if current_values else 0

            ws_power.append(styled_row(
                ws_power, ["Grupo:", "", "", "", ""]))
            ws_power.append(styled_row(
                ws_power, ["Consumo Total de Energia:", f"{total_power:.2f} W", "", "", ""]))
            ws_power.append(styled_row(
                ws_power, ["Fator de Potência Médio:", f"{avg_pf:.3f}", "", "", ""]))
            ws_power.append(styled_row(
                ws_power, ["Corrente Máxima:", f"{max_current:.2f} A", "", "", ""]))
            ws_power.append([])  # blank line

            # Electrical data header (styled)
            ws_power.append(styled_row(
                ws_power, ["Timestamp"] + electrical_names, center=True))

            # Write electrical data (no style)
            for ts in timestamps:
                row = [formatted_ts_map[ts]]
                for topic in electrical_topics:
                    row.append(sensor_lookup[topic].get(ts, 0))
                ws_power.append(row)

        # --- Save file ---
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.abspath(os.path.join(
            output_dir, f"report_{now_str}.xlsx"))
        wb.save(file_path)

        return file_path
