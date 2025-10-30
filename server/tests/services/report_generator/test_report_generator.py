import os
import pytest
from unittest.mock import MagicMock, patch
from datetime import datetime
import json
from openpyxl import load_workbook

# Importa a classe a ser testada
from services.report_generator.report_generator import ReportGenerator
from services.config_handler.config_handler import ConfigHandler


@pytest.fixture
def report_generator():
    config_handler = MagicMock()
    middleware = MagicMock()
    sensor_data_storage = MagicMock()
    return ReportGenerator(middleware, sensor_data_storage, config_handler)


def test_get_topic_mappings(report_generator):
    report_generator._config_handler.get_topic_mappings.return_value = ({
        "topic1": "name1",
        "topic2": "name2",
    }, {
        "topic1": "type1",
        "topic2": "type2",
    })

    assert report_generator._get_topic_mappings() == ({
        "topic1": "name1",
        "topic2": "name2",
    }, {
        "topic1": "type1",
        "topic2": "type2",
    })


def test_create_excel_report_fast_no_power(report_generator: ReportGenerator):
    report_generator._config_handler.get_topic_mappings.return_value = ({
        "topic1": "name1",
        "topic2": "name2",
    }, {
        "topic1": "type1",
        "topic2": "type2",
    })

    output_dir = os.path.join(os.path.dirname(__file__), "reports")

    file_path = report_generator.create_excel_report_fast({"topic1": [
        {"timestamp": "2024-01-01T10:00:00", "value": 101},
        {"timestamp": "2024-01-01T10:01:00", "value": 102},
        {"timestamp": "2024-01-01T10:02:00", "value": 103},
    ], "topic2": [
        {"timestamp": "2024-01-01T10:00:00", "value": 101},
        {"timestamp": "2024-01-01T10:01:00", "value": 102},
        {"timestamp": "2024-01-01T10:02:00", "value": 103},
    ]}, output_dir=output_dir)

    try:

        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

        # Compare Excel content instead of binary data
        wb_generated = load_workbook(file_path)
        wb_expected = load_workbook(os.path.join(
            output_dir, "no_power_example.xlsx"))

        # Check that both workbooks have the same sheet names
        assert set(wb_generated.sheetnames) == set(wb_expected.sheetnames)

        # Compare the main data sheet content
        ws_generated = wb_generated.active
        ws_expected = wb_expected.active

        # Get all values from both sheets
        generated_values = []
        expected_values = []

        for row in ws_generated.iter_rows(values_only=True):
            generated_values.append(row)

        for row in ws_expected.iter_rows(values_only=True):
            expected_values.append(row)

        # Compare the data (excluding any timestamp-based differences)
        assert len(generated_values) == len(expected_values)

        # Compare each row, allowing for timestamp differences in the filename
        for i, (gen_row, exp_row) in enumerate(zip(generated_values, expected_values)):
            if i == 0:  # Header row
                assert gen_row == exp_row
            else:
                # For data rows, compare all values except the first (timestamp)
                assert gen_row[0:] == exp_row[0:]
    finally:
        # Ensure cleanup happens even if test fails
        if os.path.exists(file_path):
            os.remove(file_path)


def test_create_excel_report_fast_no_sensor_data(report_generator: ReportGenerator):
    report_generator._config_handler.get_topic_mappings.return_value = ({
        "topic1": "name1",
        "topic2": "name2",
    }, {
        "topic1": "type1",
        "topic2": "type2",
    })
    output_dir = os.path.join(os.path.dirname(__file__), "reports")
    with pytest.raises(Exception, match="Não há dados para criar Relatório excel"):
        report_generator.create_excel_report_fast({}, output_dir=output_dir)


def test_create_excel_report_fast_power(report_generator: ReportGenerator):
    report_generator._config_handler.get_topic_mappings.return_value = ({
        "topic1": "name1",
        "topic2": "name2",
        "topic3": "name3",
        "topic4": "name4",
        "topic5": "name5",
        "topic6": "name6",
    }, {
        "topic1": "type1",
        "topic2": "type2",
        "topic3": "Current",
        "topic4": "Tension",
        "topic5": "PowerFactor",
        "topic6": "Power",
    })

    output_dir = os.path.join(os.path.dirname(__file__), "reports")

    file_path = report_generator.create_excel_report_fast({"topic1": [
        {"timestamp": "2023-01-01T10:00:00", "value": 100},
        {"timestamp": "2023-01-01T10:01:00", "value": 101},
        {"timestamp": "2023-01-01T10:02:00", "value": 102},
    ], "topic2": [
        {"timestamp": "2023-01-01T10:00:00", "value": 100},
        {"timestamp": "2023-01-01T10:01:00", "value": 101},
        {"timestamp": "2023-01-01T10:02:00", "value": 102},
    ], "topic3": [
        {"timestamp": "2024-01-01T10:00:00", "value": 100},
        {"timestamp": "2024-01-01T10:01:00", "value": 101},
        {"timestamp": "2024-01-01T10:02:00", "value": 102},
    ], "topic4": [
        {"timestamp": "2024-01-01T10:00:00", "value": 100},
        {"timestamp": "2024-01-01T10:01:00", "value": 101},
        {"timestamp": "2024-01-01T10:02:00", "value": 102},
    ], "topic5": [
        {"timestamp": "2024-01-01T10:00:00", "value": 100},
        {"timestamp": "2024-01-01T10:01:00", "value": 101},
        {"timestamp": "2024-01-01T10:02:00", "value": 102},
    ], "topic6": [
        {"timestamp": "2024-01-01T10:00:00", "value": 100},
        {"timestamp": "2024-01-01T10:01:00", "value": 101},
        {"timestamp": "2024-01-01T10:02:00", "value": 102},
    ]}, output_dir=output_dir)
    try:
        assert os.path.exists(file_path)
        assert os.path.getsize(file_path) > 0

        # Compare Excel content instead of binary data
        wb_generated = load_workbook(file_path)
        wb_expected = load_workbook(
            os.path.join(output_dir, "power_example.xlsx"))

        # Check that both workbooks have the same sheet names
        assert set(wb_generated.sheetnames) == set(wb_expected.sheetnames)

        # Compare the main data sheet content
        ws_generated = wb_generated.active
        ws_expected = wb_expected.active

        # Get all values from both sheets
        generated_values = []
        expected_values = []

        for row in ws_generated.iter_rows(values_only=True):
            generated_values.append(row)

        for row in ws_expected.iter_rows(values_only=True):
            expected_values.append(row)

        # Compare the data (excluding any timestamp-based differences)
        assert len(generated_values) == len(expected_values)

        # Compare each row, allowing for timestamp differences in the filename
        for i, (gen_row, exp_row) in enumerate(zip(generated_values, expected_values)):
            if i == 0:  # Header row
                assert gen_row == exp_row
            else:
                # For data rows, compare all values except the first (timestamp)
                assert gen_row[0:] == exp_row[0:]
    finally:
        # Ensure cleanup happens even if test fails
        if os.path.exists(file_path):
            os.remove(file_path)
